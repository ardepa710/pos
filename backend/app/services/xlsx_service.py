from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.report_service import ProductReport

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)
_ALT_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")
_LOW_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _apply_header_row(ws, headers: list[str]) -> None:
    """Write bold header row with dark-blue background into the active worksheet."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _auto_width(ws) -> None:
    """Resize each column to fit its widest cell (capped at 60)."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
            except Exception:
                cell_len = 0
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public generators
# ---------------------------------------------------------------------------


def generate_sales_xlsx(
    rows: list[dict],
    date_from: date,
    date_to: date,
) -> bytes:
    """Return Excel bytes with sales grouped by period.

    Columns: Período | No. ventas | Total MXN.
    Includes a summary row at the bottom.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Metadata rows
    ws.append(["Reporte de ventas"])
    ws.cell(row=1, column=1).font = _BOLD_FONT
    ws.append(
        [f"Del {date_from.strftime('%d/%m/%Y')} al {date_to.strftime('%d/%m/%Y')}"]
    )
    ws.append([])

    # Header row (row 4)
    headers = ["Período", "No. ventas", "Total MXN"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER

    # Data rows
    total_count = 0
    total_mxn = 0

    for data_row_idx, r in enumerate(rows, start=5):
        fill = _ALT_FILL if data_row_idx % 2 == 0 else None

        period_cell = ws.cell(row=data_row_idx, column=1, value=str(r.get("period", "")))
        count_cell = ws.cell(row=data_row_idx, column=2, value=int(r.get("count", 0)))
        mxn_cell = ws.cell(
            row=data_row_idx,
            column=3,
            value=float(r.get("total_mxn", 0)),
        )

        for cell in (period_cell, count_cell, mxn_cell):
            cell.font = _NORMAL_FONT
            if fill:
                cell.fill = fill

        count_cell.alignment = _RIGHT
        mxn_cell.alignment = _RIGHT
        mxn_cell.number_format = '"$"#,##0.00'

        total_count += int(r.get("count", 0))
        total_mxn += float(r.get("total_mxn", 0))

    # Summary row
    summary_row = (len(rows) or 0) + 5
    ws.cell(row=summary_row, column=1, value="TOTAL").font = _BOLD_FONT
    count_total = ws.cell(row=summary_row, column=2, value=total_count)
    mxn_total = ws.cell(row=summary_row, column=3, value=total_mxn)
    count_total.font = _BOLD_FONT
    count_total.alignment = _RIGHT
    mxn_total.font = _BOLD_FONT
    mxn_total.alignment = _RIGHT
    mxn_total.number_format = '"$"#,##0.00'

    _auto_width(ws)
    return _workbook_to_bytes(wb)


def generate_inventory_xlsx(rows: list[dict]) -> bytes:
    """Return Excel bytes with the current inventory snapshot.

    Columns: SKU | Nombre | Stock actual | Stock mínimo | Estado.
    Low-stock rows are highlighted in light red.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = ["SKU", "Nombre", "Stock actual", "Stock mínimo", "Unidad", "Estado"]
    _apply_header_row(ws, headers)

    for row_idx, r in enumerate(rows, start=2):
        is_low = r.get("is_low_stock", False)
        status = "BAJO" if is_low else "OK"
        fill = _LOW_FILL if is_low else (_ALT_FILL if row_idx % 2 == 0 else None)

        cells = [
            ws.cell(row=row_idx, column=1, value=r.get("sku", "")),
            ws.cell(row=row_idx, column=2, value=r.get("name", "")),
            ws.cell(row=row_idx, column=3, value=float(r.get("stock_quantity", 0))),
            ws.cell(
                row=row_idx,
                column=4,
                value=float(r.get("reorder_point", 0) or 0),
            ),
            ws.cell(row=row_idx, column=5, value=r.get("unit_of_measure", "")),
            ws.cell(row=row_idx, column=6, value=status),
        ]

        for cell in cells:
            cell.font = _BOLD_FONT if is_low else _NORMAL_FONT
            if fill:
                cell.fill = fill

        # Number format for stock columns
        ws.cell(row=row_idx, column=3).number_format = "#,##0.000"
        ws.cell(row=row_idx, column=4).number_format = "#,##0.000"
        ws.cell(row=row_idx, column=3).alignment = _RIGHT
        ws.cell(row=row_idx, column=4).alignment = _RIGHT

        # Status cell color
        if is_low:
            ws.cell(row=row_idx, column=6).font = Font(bold=True, color="DC2626", size=10)

    _auto_width(ws)
    return _workbook_to_bytes(wb)


def generate_product_report_xlsx(rows: list[ProductReport]) -> bytes:
    """Return Excel bytes with product sales figures.

    Columns: SKU | Nombre | Cantidad vendida | Ingresos MXN | Stock actual.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = ["SKU", "Nombre", "Cantidad vendida", "Ingresos MXN", "Stock actual"]
    _apply_header_row(ws, headers)

    for row_idx, r in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None

        cells = [
            ws.cell(row=row_idx, column=1, value=r.sku),
            ws.cell(row=row_idx, column=2, value=r.name),
            ws.cell(row=row_idx, column=3, value=float(r.quantity_sold)),
            ws.cell(row=row_idx, column=4, value=float(r.revenue_mxn)),
            ws.cell(row=row_idx, column=5, value=float(r.stock_current)),
        ]

        for cell in cells:
            cell.font = _NORMAL_FONT
            if fill:
                cell.fill = fill

        ws.cell(row=row_idx, column=3).number_format = "#,##0.000"
        ws.cell(row=row_idx, column=4).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=5).number_format = "#,##0.000"
        ws.cell(row=row_idx, column=3).alignment = _RIGHT
        ws.cell(row=row_idx, column=4).alignment = _RIGHT
        ws.cell(row=row_idx, column=5).alignment = _RIGHT

    _auto_width(ws)
    return _workbook_to_bytes(wb)
